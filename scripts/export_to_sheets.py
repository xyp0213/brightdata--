#!/usr/bin/env python3
"""
Export KOL Scores to Google Sheets / CSV / Excel
==================================================
Reads the scored KOL CSV/JSON and exports to:
  1) Google Sheets (optional, requires gspread + credentials)
  2) Formatted Excel (.xlsx) with auto-column-width
  3) Plain CSV (always generated)

Usage:
    python export_to_sheets.py \\
        --input data/kol_scores.csv \\
        --format google_sheets \\
        --sheet-name "KOL Rankings 2026"

Google Sheets setup:
    pip install gspread google-auth
    # Enable Google Sheets API in Google Cloud Console
    # Create service account, download credentials.json
"""

import os
import json
import csv
from datetime import datetime, timezone
from typing import Optional

import pandas as pd
from dotenv import load_dotenv

load_dotenv()


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

GOOGLE_CREDS_PATH = os.getenv("GOOGLE_SHEETS_CREDENTIALS_PATH", "credentials.json")
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "")


# ---------------------------------------------------------------------------
# Export: CSV
# ---------------------------------------------------------------------------

def export_csv(df: pd.DataFrame, path: str) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  CSV exported → {path}")
    return path


# ---------------------------------------------------------------------------
# Export: Excel (formatted)
# ---------------------------------------------------------------------------

def export_excel(df: pd.DataFrame, path: str, sheet_name: str = "KOL Rankings") -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

        # Format the sheet
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header styling
        header_font = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1, value=f"KOL Scoring Report — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        title_cell.font = Font(name="Segoe UI", bold=True, size=12, color="1E293B")
        title_cell.alignment = Alignment(horizontal="left")

        # Style header row (row 2)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Style data rows
        data_font = Font(name="Segoe UI", size=10)
        for row_idx in range(3, len(df) + 3):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # Highlight top 3 rows
            if row_idx <= 5:
                row_fill = PatternFill(
                    start_color=("FEF3C7" if row_idx == 3 else "F0F9FF" if row_idx == 4 else "F8FAFC"),
                    end_color=("FEF3C7" if row_idx == 3 else "F0F9FF" if row_idx == 4 else "F8FAFC"),
                    fill_type="solid",
                )
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill

        # Auto-fit column widths
        for col_idx in range(1, len(df.columns) + 1):
            max_len = max(
                len(str(ws.cell(row=2, column=col_idx).value or "")),
                max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(3, min(len(df) + 3, 53))),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)

    print(f"  Excel exported → {path}")
    return path


# ---------------------------------------------------------------------------
# Export: Google Sheets
# ---------------------------------------------------------------------------

def export_google_sheets(
    df: pd.DataFrame,
    sheet_id: Optional[str] = None,
    sheet_name: str = "KOL Rankings",
    credentials_path: Optional[str] = None,
) -> str:
    """
    Export DataFrame to a Google Sheet.
    Requires service account credentials with Sheets API access.
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise ImportError(
            "gspread and google-auth are required for Google Sheets export. "
            "Install with: pip install gspread google-auth"
        )

    creds_path = credentials_path or GOOGLE_CREDS_PATH
    gsheet_id = sheet_id or GOOGLE_SHEET_ID

    if not os.path.exists(creds_path):
        raise FileNotFoundError(
            f"Google credentials not found at {creds_path}. "
            "Download from Google Cloud Console → APIs & Services → Credentials."
        )
    if not gsheet_id:
        raise ValueError(
            "GOOGLE_SHEET_ID not set. Find it in your sheet URL: "
            "https://docs.google.com/spreadsheets/d/<SHEET_ID>/edit"
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    client = gspread.authorize(creds)

    # Open or create sheet
    try:
        spreadsheet = client.open_by_key(gsheet_id)
    except gspread.exceptions.SpreadsheetNotFound:
        spreadsheet = client.create(f"KOL Rankings — {datetime.now().strftime('%Y%m%d')}")
        # Share with service account email is automatic for service accounts
        print(f"  Created new spreadsheet: {spreadsheet.url}")

    # Select or create worksheet
    try:
        ws = spreadsheet.worksheet(sheet_name)
        ws.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=30)

    # Replace NaN/Inf
    df_clean = df.fillna("").replace([float("inf"), float("-inf")], "")

    # Upload data
    data = [df_clean.columns.tolist()] + df_clean.values.tolist()
    ws.update("A1", data, value_input_option="USER_ENTERED")

    # Format header row
    ws.format("A1:Z1", {
        "textFormat": {"bold": True, "fontSize": 10},
        "backgroundColor": {"red": 0.145, "green": 0.388, "blue": 0.890},  # #2563EB
        "horizontalAlignment": "CENTER",
    })

    print(f"  Google Sheets exported → {spreadsheet.url}")
    return spreadsheet.url


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Export KOL scores to CSV, Excel, or Google Sheets"
    )
    parser.add_argument(
        "--input", "-i", required=True,
        help="Input CSV or JSON file from kol_scoring_model.py"
    )
    parser.add_argument(
        "--format", "-f", choices=["csv", "excel", "google_sheets", "all"],
        default="excel",
        help="Export format(s)"
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output path (for csv/excel)"
    )
    parser.add_argument(
        "--sheet-name", default="KOL Rankings",
        help="Sheet/worksheet name"
    )
    parser.add_argument(
        "--sheet-id",
        help="Google Sheet ID (overrides .env)"
    )
    parser.add_argument(
        "--credentials",
        help="Path to Google service account credentials JSON"
    )
    args = parser.parse_args()

    # Load input
    if args.input.endswith(".json"):
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
        df = pd.DataFrame(data)
    else:
        df = pd.read_csv(args.input)

    print(f"\nLoaded {len(df)} records from {args.input}\n")

    export_format = args.format

    # CSV
    if export_format in ("csv", "all"):
        out = args.output or args.input.replace(".csv", "_exported.csv")
        if not out.endswith(".csv"):
            out += ".csv"
        export_csv(df, out)

    # Excel
    if export_format in ("excel", "all"):
        out = args.output or args.input.replace(".csv", "_formatted.xlsx")
        if not out.endswith(".xlsx"):
            out += ".xlsx"
        export_excel(df, out, sheet_name=args.sheet_name)

    # Google Sheets
    if export_format in ("google_sheets", "all"):
        url = export_google_sheets(
            df,
            sheet_id=args.sheet_id,
            sheet_name=args.sheet_name,
            credentials_path=args.credentials,
        )
        print(f"\n  Open in browser: {url}")
